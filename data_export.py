"""
data_export.py — Экспорт данных платформы в форматы CSV, Excel, JSON.
═══════════════════════════════════════════════════════════════════════════════
Для обработки в R, SPSS, Excel, LaTeX и других инструментах.

Форматы:
  1. CSV с заголовками (UTF-8, разделитель ;)
  2. Excel (.xlsx) с листами по разделам
  3. JSON структурированный
  4. LaTeX-таблицы (для вставки в диссертацию)
═══════════════════════════════════════════════════════════════════════════════
"""
from __future__ import annotations

import os
import csv
import json
from typing import List, Dict, Optional

import numpy as np

_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
_EXPORT_DIR = os.path.join(_DATA_DIR, "export")
os.makedirs(_EXPORT_DIR, exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════
# 1. ЭКСПОРТ В CSV
# ═══════════════════════════════════════════════════════════════════════════
def export_csv(data: List[Dict], filename: str,
               delimiter: str = ";") -> str:
    """Экспорт списка словарей в CSV."""
    if not data:
        return ""
    path = os.path.join(_EXPORT_DIR, filename)
    keys = list(data[0].keys())
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=keys, delimiter=delimiter)
        writer.writeheader()
        for row in data:
            writer.writerow({k: row.get(k, "") for k in keys})
    return path


def export_timeseries_csv(times: List, values: List,
                          series_name: str = "value",
                          filename: str = "timeseries.csv") -> str:
    """Экспорт временного ряда в CSV."""
    data = [{"time_min": t, series_name: v} for t, v in zip(times, values)]
    return export_csv(data, filename)


def export_case_base_csv(cases, filename: str = "case_base.csv") -> str:
    """Экспорт базы прецедентов в CSV."""
    from cbr_engine import FEATURE_NAMES
    data = []
    for case in cases:
        row = {
            "case_id": case.case_id,
            "source_file": case.source_file,
            "fire_rank": case.fire_rank,
            "fuel_type": case.fuel_type,
            "roof_type": case.roof_type,
            "duration_min": case.duration_min,
            "localized": case.localized,
            "extinguished": case.extinguished,
            "foam_attacks": case.foam_attacks,
            "foam_success": case.foam_success,
            "n_incidents": case.n_incidents,
            "cluster_id": case.cluster_id,
            "cluster_name": case.cluster_name,
            "situation_type": case.situation_type,
        }
        for i, name in enumerate(FEATURE_NAMES):
            row[name] = case.features[i] if i < len(case.features) else 0
        data.append(row)
    return export_csv(data, filename)


# ═══════════════════════════════════════════════════════════════════════════
# 2. ЭКСПОРТ В EXCEL
# ═══════════════════════════════════════════════════════════════════════════
def export_excel(sheets: Dict[str, List[Dict]],
                 filename: str = "saur_data.xlsx") -> str:
    """Экспорт в Excel с несколькими листами.

    sheets: {"Прецеденты": [rows], "Статистика": [rows], ...}
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        # Fallback: сохранить каждый лист как отдельный CSV
        paths = []
        for sheet_name, data in sheets.items():
            safe_name = sheet_name.replace(" ", "_")[:20]
            p = export_csv(data, f"{safe_name}.csv")
            paths.append(p)
        return "; ".join(paths)

    path = os.path.join(_EXPORT_DIR, filename)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, data in sheets.items():
        if not data:
            continue
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit
        keys = list(data[0].keys())
        # Заголовки
        for col, key in enumerate(keys, 1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.font = cell.font.copy(bold=True)
        # Данные
        for row_idx, row in enumerate(data, 2):
            for col, key in enumerate(keys, 1):
                val = row.get(key, "")
                if isinstance(val, (list, dict)):
                    val = str(val)
                ws.cell(row=row_idx, column=col, value=val)

    wb.save(path)
    return path


# ═══════════════════════════════════════════════════════════════════════════
# 3. ЭКСПОРТ В LATEX
# ═══════════════════════════════════════════════════════════════════════════
def export_latex_table(headers: List[str], rows: List[List],
                       caption: str = "", label: str = "",
                       filename: str = "table.tex") -> str:
    """Генерация LaTeX-таблицы для вставки в диссертацию."""
    path = os.path.join(_EXPORT_DIR, filename)
    n_cols = len(headers)
    col_spec = "|" + "|".join(["l"] * n_cols) + "|"

    lines = []
    lines.append("\\begin{table}[htbp]")
    lines.append("\\centering")
    if caption:
        lines.append(f"\\caption{{{caption}}}")
    if label:
        lines.append(f"\\label{{{label}}}")
    lines.append(f"\\begin{{tabular}}{{{col_spec}}}")
    lines.append("\\hline")
    lines.append(" & ".join(f"\\textbf{{{h}}}" for h in headers) + " \\\\")
    lines.append("\\hline")
    for row in rows:
        cells = [str(v) for v in row]
        lines.append(" & ".join(cells) + " \\\\")
    lines.append("\\hline")
    lines.append("\\end{tabular}")
    lines.append("\\end{table}")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


# ═══════════════════════════════════════════════════════════════════════════
# 4. ПОЛНЫЙ ЭКСПОРТ ПЛАТФОРМЫ
# ═══════════════════════════════════════════════════════════════════════════
def full_export(
    case_base=None,
    stat_results: Optional[Dict] = None,
    autonomy_results: Optional[Dict] = None,
    adaptation_metrics: Optional[Dict] = None,
) -> Dict[str, str]:
    """Полный экспорт всех данных платформы.

    Возвращает: {"csv_cases": path, "excel": path, "latex": path, ...}
    """
    paths = {}

    # CSV: база прецедентов
    if case_base and hasattr(case_base, 'cases') and case_base.cases:
        paths["csv_cases"] = export_case_base_csv(case_base.cases)

    # CSV: статистика
    if stat_results:
        if "descriptive" in stat_results:
            paths["csv_descriptive"] = export_csv(
                stat_results["descriptive"], "descriptive_stats.csv")
        if "significant_correlations" in stat_results:
            paths["csv_correlations"] = export_csv(
                stat_results["significant_correlations"], "correlations.csv")
        if "anova" in stat_results:
            paths["csv_anova"] = export_csv(
                stat_results["anova"], "anova_results.csv")

    # Excel: всё в одном файле
    sheets = {}
    if stat_results and "descriptive" in stat_results:
        sheets["Описательная"] = stat_results["descriptive"]
    if stat_results and "significant_correlations" in stat_results:
        sheets["Корреляции"] = stat_results["significant_correlations"]
    if stat_results and "anova" in stat_results:
        sheets["Дисперсионный"] = stat_results["anova"]
    if adaptation_metrics:
        sheets["Адаптация"] = [adaptation_metrics]
    if sheets:
        paths["excel"] = export_excel(sheets, "saur_full_export.xlsx")

    # LaTeX: ключевые таблицы для диссертации
    if stat_results and "descriptive" in stat_results:
        desc = stat_results["descriptive"]
        rows = [[d["name"][:25], d["n"], d["M"], d["SD"], d["CI95"]]
                for d in desc[:8]]
        paths["latex_descriptive"] = export_latex_table(
            ["Признак", "n", "M", "SD", "95\\% ДИ"],
            rows,
            caption="Описательная статистика признаков пожаров",
            label="tab:descriptive",
            filename="tab_descriptive.tex")

    return paths


if __name__ == "__main__":
    from cbr_engine import generate_demo_casebase
    from stat_analysis import full_analysis

    cb = generate_demo_casebase(50)
    stats = full_analysis(cb)

    paths = full_export(case_base=cb, stat_results=stats)
    print("Экспорт данных:")
    for name, path in paths.items():
        print(f"  {name}: {path}")
